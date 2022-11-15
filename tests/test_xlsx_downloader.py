from itertools import chain
from pathlib import Path
from unittest.mock import MagicMock

from openpyxl import load_workbook

from ckanext.versioned_datastore.lib.downloads.xlsx import (
    xlsx_writer,
    ALL_IN_ONE_FILE_NAME,
    DEFAULT_SHEET_NAME,
    RESOURCE_ID_FIELD_NAME,
)


def test_xlsx_single(tmpdir: Path):
    data = {'field1': 'A', 'field2': 'B', 'field3': 'C'}
    resource_id = 'resource-a'
    field_counts = {resource_id: {field: 10 for field in data}}
    request = MagicMock(separate_files=False, ignore_empty_fields=False)
    hit = MagicMock()

    with xlsx_writer(request, tmpdir, field_counts) as write:
        write(hit, data, resource_id)

    data_file = tmpdir / ALL_IN_ONE_FILE_NAME
    assert data_file.exists()

    workbook = load_workbook(filename=data_file)
    rows = list(workbook[DEFAULT_SHEET_NAME].rows)
    assert len(rows) == 2

    header = list(cell.value for cell in rows[0])
    expected_headers = [RESOURCE_ID_FIELD_NAME] + sorted(data)
    assert header == expected_headers

    row = rows[1]
    row_data = dict(zip(header, [cell.value for cell in row]))
    assert row_data.pop(RESOURCE_ID_FIELD_NAME) == resource_id
    assert row_data == data


def test_xlsx_multiple_one_file(tmpdir: Path):
    data_a = {'field1': 'A', 'field2': 'B', 'field3': 'C'}
    data_b = {'field5': 'A', 'field1': 'B', 'field7': 'C'}
    resource_id_a = 'resource-a'
    resource_id_b = 'resource-b'
    field_counts = {
        resource_id_a: {field: 10 for field in data_a},
        resource_id_b: {field: 10 for field in data_b},
    }
    request = MagicMock(separate_files=False, ignore_empty_fields=False)
    hit = MagicMock()

    with xlsx_writer(request, tmpdir, field_counts) as write:
        write(hit, data_a, resource_id_a)
        write(hit, data_b, resource_id_b)

    data_file = tmpdir / ALL_IN_ONE_FILE_NAME
    assert data_file.exists()

    workbook = load_workbook(filename=data_file)
    rows = list(workbook[DEFAULT_SHEET_NAME].rows)
    header = list(cell.value for cell in rows[0])
    expected_headers = [RESOURCE_ID_FIELD_NAME] + sorted(set(chain(data_a, data_b)))
    assert header == expected_headers

    row_a = rows[1]
    row_a_data = dict(zip(header, [cell.value for cell in row_a]))
    assert row_a_data.pop(RESOURCE_ID_FIELD_NAME) == resource_id_a
    assert row_a_data.pop('field5') is None
    assert row_a_data.pop('field7') is None
    assert row_a_data == data_a

    row_b = rows[2]
    row_b_data = dict(zip(header, [cell.value for cell in row_b]))
    assert row_b_data.pop(RESOURCE_ID_FIELD_NAME) == resource_id_b
    assert row_b_data.pop('field2') is None
    assert row_b_data.pop('field3') is None
    assert row_b_data == data_b


def test_xlsx_multiple_separate_file(tmpdir: Path):
    data_a = {'field1': 'A', 'field2': 'B', 'field3': 'C'}
    data_b = {'field5': 'A', 'field1': 'B', 'field7': 'C'}
    resource_id_a = 'resource-a'
    resource_id_b = 'resource-b'
    field_counts = {
        resource_id_a: {field: 10 for field in data_a},
        resource_id_b: {field: 10 for field in data_b},
    }
    request = MagicMock(separate_files=True, ignore_empty_fields=False)
    hit = MagicMock()

    with xlsx_writer(request, tmpdir, field_counts) as write:
        write(hit, data_a, resource_id_a)
        write(hit, data_b, resource_id_b)

    for resource_id, data in [(resource_id_a, data_a), (resource_id_b, data_b)]:
        data_file = tmpdir / f'{resource_id}.xlsx'
        assert data_file.exists()

        workbook = load_workbook(filename=data_file)
        rows = list(workbook[DEFAULT_SHEET_NAME].rows)
        assert len(rows) == 2

        header = list(cell.value for cell in rows[0])
        expected_headers = sorted(data)
        assert header == expected_headers

        row = rows[1]
        row_data = dict(zip(header, [cell.value for cell in row]))
        assert row_data == data
