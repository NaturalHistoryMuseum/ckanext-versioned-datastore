from openpyxl import Workbook, load_workbook

from .base import BaseDerivativeGenerator
from ..utils import flatten_dict


class XlsxDerivativeGenerator(BaseDerivativeGenerator):
    name = 'xlsx'
    extension = 'xlsx'

    DEFAULT_SHEET_NAME = 'Data'

    def __init__(self, output_dir, fields, query, resource_id=None, **format_args):
        super(XlsxDerivativeGenerator, self).__init__(
            output_dir, fields, query, resource_id, **format_args
        )
        self.workbook = None

    def initialise(self):
        self.workbook.create_sheet(self.DEFAULT_SHEET_NAME)
        self.workbook.active.append(self.fields['main'])
        super(XlsxDerivativeGenerator, self).initialise()

    def setup(self):
        try:
            self.workbook = load_workbook(self.file_paths['main'])
        except Exception as e:
            self.workbook = Workbook(write_only=True)
        super(XlsxDerivativeGenerator, self).setup()

    def finalise(self):
        try:
            self.workbook.save(self.file_paths['main'])
        finally:
            # if something goes wrong when trying to save the workbook, make sure to
            # close the workbook before raising the error
            self.workbook.close()
        super(XlsxDerivativeGenerator, self).finalise()

    def _write(self, record):
        row = flatten_dict(record)
        filtered_row = {}
        for field, value in row.items():
            if value is None and field not in self.fields['main']:
                continue
            elif field not in self.fields['main']:
                raise ValueError('Unexpected field.')
            else:
                filtered_row[field] = value
        if self.resource_id:
            filtered_row[self.RESOURCE_ID_FIELD_NAME] = self.resource_id
        self.workbook.active.append(
            filtered_row.get(field) for field in self.fields['main']
        )
