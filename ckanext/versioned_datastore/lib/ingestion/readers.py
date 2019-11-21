import numbers

import abc
import openpyxl
import six
import unicodecsv
import xlrd
from cchardet import UniversalDetector
from ckanext.versioned_datastore.lib.ingestion import exceptions
from openpyxl.cell.read_only import EmptyCell

from . import utils
from .utils import ensure_reset, iter_universal_lines


def get_reader(resource_format):
    '''
    Given a format, return a ResourceReader instance for it. If no ResourceReader type cannot be
    matched, None is returned.

    :param resource_format: the format
    :return: a ResourceReader instance or None
    '''
    resource_format = resource_format.lower()
    if resource_format in utils.CSV_FORMATS:
        return SVReader(u'excel')
    if resource_format in utils.TSV_FORMATS:
        return SVReader(u'excel-tab')
    if resource_format in utils.XLS_FORMATS:
        return XLSReader()
    if resource_format in utils.XLSX_FORMATS:
        return XLSXReader()

    return None


@six.add_metaclass(abc.ABCMeta)
class ResourceReader(object):
    '''
    Abstract class to read fields and rows from a resource.
    '''

    def __init__(self, compressible):
        '''
        :param compressible: whether the reader can cope with a gzipped file pointer being passed to
                             the get_fields and get_rows functions.
        '''
        self.compressible = compressible

    @abc.abstractmethod
    def get_fields(self, resource_data_fp):
        '''
        Returns a list of field names in the order they were provided in the source resource.

        :param resource_data_fp: the file pointer to the source
        :return: a list of field names
        '''
        pass

    @abc.abstractmethod
    def _get_rows(self, resource_data_fp):
        '''
        Returns a generator which yields rows as dicts.

        :param resource_data_fp: the file pointer to the source
        :return: a generator of row dicts
        '''
        pass

    def modify_metadata(self, metadata):
        '''
        Allows for modification of the metadata before it is written out in the intermediate file.
        The metadata should be modified in place. By default this function does nothing and should
        be overridden by subclasses.

        :param metadata: the current metadata as a dict
        '''
        pass

    def iter_rows(self, resource_data_fp):
        '''
        Returns a generator which yields the rows as dicts. Each row is checked for an _id field
        and if one is found its value is converted to an int.

        :param resource_data_fp: the file pointer to the source
        :return: a generator of row dicts
        '''
        for row_number, row in enumerate(self._get_rows(resource_data_fp)):
            if u'_id' in row:
                try:
                    row[u'_id'] = int(row[u'_id'])
                except ValueError as e:
                    raise exceptions.InvalidId(row_number, row, cause=e)
            yield row


class SVReader(ResourceReader):
    '''
    A *SV reader - handles CSVs and TSVs.
    '''

    def __init__(self, dialect):
        '''
        :param dialect: the dialect of the source, this is passed straight to the csv reader
                        constructor function
        '''
        super(SVReader, self).__init__(True)
        self.dialect = dialect
        self.encoding = None

    def _get_dict_reader(self, resource_data_fp):
        '''
        Returns a dict reader for the given source file pointer. If the encoding of the source
        hasn't been guessed yet then this function will read the from the file pointer until EOF and
        guess the encoding. The file pointer is reset to the start after this occurs.

        :param resource_data_fp: the file pointer to the source
        :return: a DictReader object
        '''
        if self.encoding is None:
            with ensure_reset(resource_data_fp):
                detector = UniversalDetector()
                while True:
                    chunk = resource_data_fp.read(8192)
                    if chunk:
                        detector.feed(chunk)
                    else:
                        detector.close()
                        break

            self.encoding = detector.result[u'encoding']
            # if the detector failed to work out the encoding (unlikely) or if the encoding it
            # comes up with is ASCII, just default to UTF-8 (UTF-8 is a superset of ASCII)
            if self.encoding is None or self.encoding == u'ASCII':
                self.encoding = u'utf-8'

        # create and return the dict reader
        line_iterator = iter_universal_lines(resource_data_fp, self.encoding)
        return unicodecsv.DictReader(line_iterator, dialect=self.dialect, encoding=self.encoding)

    def _get_rows(self, resource_data_fp):
        '''
        Returns a generator which yields rows as dicts using a DictReader. If the encoding of the
        source hasn't been guessed yet then this function will read the from the file pointer until
        EOF and guess the encoding. The file pointer is reset to the start after this occurs.

        :param resource_data_fp: the file pointer to the source
        :return: a generator of row dicts
        '''
        with ensure_reset(resource_data_fp):
            reader = self._get_dict_reader(resource_data_fp)
            for row in reader:
                yield row

    def get_fields(self, resource_data_fp):
        '''
        Returns a list of field names in the order they were provided in the source resource. This
        is achieved by reading the first line from the given file pointer.

        :param resource_data_fp: the file pointer to the source
        :return: a list of field names
        '''
        with ensure_reset(resource_data_fp):
            reader = self._get_dict_reader(resource_data_fp)
            return reader.unicode_fieldnames

    def modify_metadata(self, metadata):
        '''
        Modify the metadata to include the original encoding of the file we've read.

        :param metadata: the current metadata as a dict
        '''
        metadata[u'original_encoding'] = self.encoding


class XLSReader(ResourceReader):
    '''
    Old style Excel file reader.
    '''

    def __init__(self):
        # compressible needs to be false here as xlrd opens the file using it's name, not the byte
        # stream
        super(XLSReader, self).__init__(False)

    def _iter_rows(self, resource_data_fp):
        '''
        Returns a generator of rows from the resource file pointer. Each row is a list of cells.

        :param resource_data_fp: the file pointer to the source
        :return: a generator of cell lists
        '''
        with ensure_reset(resource_data_fp):
            # open the xls file up
            book = xlrd.open_workbook(resource_data_fp.name)
            # select the first sheet by default
            sheet = book.sheet_by_index(0)
            # get_rows is a generator so just return it
            return sheet.get_rows()

    def _get_rows(self, resource_data_fp):
        '''
        Returns a generator of row dicts from the source.

        :param resource_data_fp: the file pointer to the source
        :return: a generator of row dicts
        '''
        rows = self._iter_rows(resource_data_fp)
        # assume the first row is the header
        header = [unicode(cell.value) for cell in next(rows)]
        # then read all the other rows as data
        for row in rows:
            data = {}
            for field, cell in zip(header, row):
                # if the cell is the id column, it contains a number and the number is an
                # integer, convert it from a float to an int
                if field == u'_id':
                    if cell.ctype == xlrd.XL_CELL_NUMBER and cell.value.is_integer():
                        data[field] = int(cell.value)
                    else:
                        raise Exception(u'_id not int')
                elif cell == xlrd.XL_CELL_EMPTY:
                    # ignore empty cells
                    continue
                else:
                    # otherwise just use the value (if it's not an empty string)
                    value = unicode(cell.value)
                    if value:
                        data[field] = unicode(cell.value)
            # yield the row
            yield data

    def get_fields(self, resource_data_fp):
        '''
        Returns a list of field names in the order they were provided in the source resource. This
        is achieved by assuming the first line in the worksheet is the header row.

        :param resource_data_fp: the file pointer to the source
        :return: a list of field names
        '''
        # assume the first row is the header
        return [unicode(cell.value) for cell in next(self._iter_rows(resource_data_fp))]


class XLSXReader(ResourceReader):
    '''
    New style Excel file reader.
    '''

    def __init__(self):
        # compressible needs to be false here as openpyxl can't handle it
        super(XLSXReader, self).__init__(False)

    def _iter_rows(self, resource_data_fp):
        '''
        Returns a generator of rows from the resource file pointer. Each row is a list of cells.

        :param resource_data_fp: the file pointer to the source
        :return: a generator of cell lists
        '''
        with ensure_reset(resource_data_fp):
            workbook = openpyxl.load_workbook(resource_data_fp, read_only=True)
            # get a generator for the rows in the first worksheet and return it
            return workbook.worksheets[0].iter_rows()

    def _get_rows(self, resource_data_fp):
        '''
        Returns a generator of row dicts from the source.

        :param resource_data_fp: the file pointer to the source
        :return: a generator of row dicts
        '''
        # get a generator for the rows in the first worksheet
        rows = self._iter_rows(resource_data_fp)
        # assume the first row is a header
        header = [unicode(cell.value) for cell in next(rows)]
        # then read all the other rows as data
        for row in rows:
            data = {}
            for field, cell in zip(header, row):
                # if the cell is the id column and it contains a number make sure it stays a
                # number
                if field == u'_id':
                    if isinstance(cell.value, numbers.Integral):
                        data[field] = cell.value
                    else:
                        raise Exception(u'_id not int')
                # ignore empty cells
                elif isinstance(cell, EmptyCell):
                    continue
                else:
                    # convert everything else to unicode
                    value = unicode(cell.value)
                    if value:
                        data[field] = value
            # yield the row
            yield data

    def get_fields(self, resource_data_fp):
        '''
        Returns a list of field names in the order they were provided in the source resource. This
        is achieved by assuming the first line in the worksheet is the header row.

        :param resource_data_fp: the file pointer to the source
        :return: a list of field names
        '''
        # assume the first row is the header
        return [unicode(cell.value) for cell in next(self._iter_rows(resource_data_fp))]


class APIReader(ResourceReader):
    '''
    Direct API data upload reader. The data comes to us as a list of dicts.
    '''

    def __init__(self, data):
        '''
        :param data: the data as a list of dicts
        '''
        super(APIReader, self).__init__(False)
        self.data = data

    def _get_rows(self, *args, **kwargs):
        '''
        Returns a generator of row dicts from the sent data, this just returns an iterator for the
        data list.

        :return: a generator of row dicts
        '''
        return iter(self.data)

    def get_fields(self, *args, **kwargs):
        '''
        Returns a list of field names. We can't really guarantee order here because python2 doesn't
        guarantee dict key ordering :(

        :return: a list of field names
        '''
        fields = []
        for row in self.data:
            for field in row.keys():
                if field not in fields:
                    fields.append(field)
        return fields
